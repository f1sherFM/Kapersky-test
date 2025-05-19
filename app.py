import pandas as pd
import requests
from bs4 import BeautifulSoup
import difflib
import re
import os
from datetime import datetime
import json
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from tenacity import retry, stop_after_attempt, wait_exponential
from tqdm import tqdm
from html_report import generate_html_report

# --- Конфигурация ---
INPUT_CSV = r".\Test_check.csv"
OUTPUT_DIR = "results"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# --- Функции ---

def ensure_directory_exists():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def extract_text_from_url(URL):
    """Извлекает основной текст статьи по URL"""
    try:
        response = requests.get(URL, headers=HEADERS, timeout=20)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Удаляем ненужные элементы
        for tag in ['script', 'style', 'nav', 'footer', 'iframe', 'img', 'button']:
            for element in soup.find_all(tag):
                element.decompose()

        # Поиск основного контента
        article_body = (
            soup.find('article') or
            soup.find(class_=re.compile(r'(post-body|article-text|entry-content|articleBody)', re.I)) or
            soup.find(itemprop='articleBody') or
            soup.find('div', class_='article__text') or
            soup.find('body')
        )

        text = article_body.get_text(separator='\n', strip=True) if article_body else soup.get_text(separator='\n', strip=True)
        return re.sub(r'\n{2,}', '\n\n', text).strip()

    except Exception as e:
        print(f"❌ Ошибка при загрузке {URL}: {e}")
        return f"Ошибка: {str(e)}"

def compare_texts(original_text, lib_text):
    """Сравнивает два текста и возвращает показатели"""

    def normalize(text):
        return re.sub(r'\s+', ' ', text.strip().lower())

    matcher = difflib.SequenceMatcher(None, normalize(original_text), normalize(lib_text))
    similarity = matcher.ratio() * 100  # Процент сходства

    original_lines = [line.strip() for line in original_text.split('\n') if line.strip()]
    lib_lines = [line.strip() for line in lib_text.split('\n') if line.strip()]

    missing = [line for line in original_lines if normalize(line) not in normalize(lib_text)]
    extra = [line for line in lib_lines if normalize(line) not in normalize(original_text)]

    return {
        'similarity': round(similarity, 2),
        'original_length': len(original_text),
        'lib_length': len(lib_text),
        'missing_lines_count': len(missing),
        'extra_lines_count': len(extra),
        'example_missing': list(set(missing))[:3],
        'example_extra': list(set(extra))[:3]
    }

def generate_ai_analysis(comparison_data):
    """Генерирует аналитический отчет на основе сравнения текстов"""
    analysis = []
    
    # Анализ сходства
    similarity = comparison_data['similarity']
    if similarity > 80:
        analysis.append("✅ Тексты имеют высокое сходство (более 80%). Основное содержание сохранено.")
    elif similarity > 50:
        analysis.append("⚠️ Тексты имеют умеренное сходство (50-80%). Имеются существенные расхождения.")
    else:
        analysis.append("❌ Тексты имеют низкое сходство (менее 50%). Контент значительно отличается.")

    # Анализ пропущенных строк
    missing_count = comparison_data['missing_lines_count']
    if missing_count > 0:
        analysis.append(f"\n🔍 Пропущено строк: {missing_count}")
        analysis.append("Наиболее значимые пропуски:")
        for example in comparison_data['example_missing'][:3]:
            analysis.append(f"- {example[:150]}{'...' if len(example) > 150 else ''}")
    
    # Анализ добавленных строк
    extra_count = comparison_data['extra_lines_count']
    if extra_count > 0:
        analysis.append(f"\n🔍 Добавлено строк: {extra_count}")
        analysis.append("Наиболее значимые добавления:")
        for example in comparison_data['example_extra'][:3]:
            analysis.append(f"- {example[:150]}{'...' if len(example) > 150 else ''}")

    # Общий вывод
    analysis.append("\n📌 Вывод:")
    if missing_count > extra_count:
        analysis.append("Основное расхождение - пропуск значительных частей оригинального текста.")
    elif extra_count > missing_count:
        analysis.append("Основное расхождение - добавление нового контента, отсутствующего в оригинале.")
    else:
        analysis.append("Расхождения включают как пропуски, так и добавления контента.")

    return "\n".join(analysis)

def save_results(results):
    """Сохраняет результаты в JSON и CSV"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    json_path = os.path.join(OUTPUT_DIR, f'results_{timestamp}.json')
    
    # Формируем полный отчет с AI-анализом
    full_report = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "source_csv": INPUT_CSV,
            "total_articles": len(results),
            "successful": len([r for r in results if r['status'] == 'success']),
            "failed": len([r for r in results if r['status'] == 'error'])
        },
        "articles": []
    }
    
    for result in results:
        article_data = {
            "url": result['url'],
            "status": result['status'],
            "timestamp": datetime.now().isoformat()
        }
        
        if result['status'] == 'success':
            article_data.update({
                "similarity": result['similarity'],
                "length_analysis": {
                    "original": result['original_length'],
                    "lib": result['lib_length'],
                    "difference": abs(result['original_length'] - result['lib_length'])
                },
                "content_differences": {
                    "missing_lines": result['missing_lines_count'],
                    "extra_lines": result['extra_lines_count'],
                    "example_missing": result['example_missing'],
                    "example_extra": result['example_extra']
                },
                "ai_analysis": generate_ai_analysis(result)
            })
        else:
            article_data['error'] = result['error']
        
        full_report["articles"].append(article_data)
    
    # Сохраняем JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(full_report, f, ensure_ascii=False, indent=2, default=str)
    
    # Сохраняем CSV (упрощенная версия)
    csv_path = os.path.join(OUTPUT_DIR, f'results_{timestamp}.csv')
    pd.DataFrame(results).to_csv(csv_path, index=False, encoding='utf-8-sig')
    
    print(f"\n✅ Результаты сохранены:\n- Полный отчет: {json_path}\n- Таблица данных: {csv_path}")

def generate_comprehensive_report(results):
    """Генерирует комплексный отчет в формате Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = os.path.join(OUTPUT_DIR, f'comprehensive_report_{timestamp}.xlsx')
    
    # Создаем книгу Excel
    wb = Workbook()
    
    # === Лист 1: Сводка ===
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    
    # Заголовки
    headers = [
        "URL", "Сходство (%)", "Статус",
        "Длина эталона", "Длина извлеченного", "Разница",
        "Пропущено строк", "Добавлено строк", "Оценка значимости пропусков",
        "Комментарий (советы по улучшению)"  # Новый столбец
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws_summary.cell(row=1, column=col_num, value=header).font = Font(bold=True)
    
    # Данные
    for row_num, result in enumerate(results, 2):
        ws_summary.cell(row=row_num, column=1, value=result['url'])
        
        if result['status'] == 'success':
            ws_summary.cell(row=row_num, column=2, value=result['similarity'])
            ws_summary.cell(row=row_num, column=4, value=result['original_length'])
            ws_summary.cell(row=row_num, column=5, value=result['lib_length'])
            ws_summary.cell(row=row_num, column=6, value=abs(result['original_length'] - result['lib_length']))
            ws_summary.cell(row=row_num, column=7, value=result['missing_lines_count'])
            ws_summary.cell(row=row_num, column=8, value=result['extra_lines_count'])
            
            # Оценка значимости пропусков
            missing_importance = "Низкая"
            if result['missing_lines_count'] > 10:
                missing_importance = "Средняя"
            if result['missing_lines_count'] > 30 or any(len(ex) > 100 for ex in result['example_missing']):
                missing_importance = "Высокая"
            ws_summary.cell(row=row_num, column=9, value=missing_importance)
            
            # Генерация комментариев для улучшения (новый столбец)
            comments = []
            
            # Советы по пропущенному контенту
            if result['missing_lines_count'] > 0:
                if any(re.search(r'(адрес|контакт|телефон)', ex, re.I) for ex in result['example_missing']):
                    comments.append("Добавить обработку контактных данных.")
                if any(re.search(r'(вывод|заключение|итог)', ex, re.I) for ex in result['example_missing']):
                    comments.append("Улучшить извлечение ключевых выводов.")
                if result['missing_lines_count'] > 15:
                    comments.append("Расширить правила парсинга для этого типа контента.")
            
            # Советы по добавленному контенту
            if result['extra_lines_count'] > 0:
                if any(re.search(r'(реклам|промо|акци)', ex, re.I) for ex in result['example_extra']):
                    comments.append("Добавить фильтрацию рекламного контента.")
                if result['extra_lines_count'] > 10:
                    comments.append("Проверить правила исключения нерелевантных блоков.")
            
            # Общие советы
            if result['similarity'] < 50:
                comments.append("Требуется глубокая оптимизация парсинга.")
            elif result['similarity'] < 70:
                comments.append("Рекомендуется донастройка правил извлечения.")
            
            # Если нет особых проблем
            if not comments and result['similarity'] > 80:
                comments.append("Качество извлечения хорошее. Можно оптимизировать обработку специфичных элементов.")
            
            ws_summary.cell(row=row_num, column=10, value="\n".join(comments) if comments else "Не требует значительных улучшений")
        else:
            ws_summary.cell(row=row_num, column=2, value="N/A")
            error = result.get('error', 'Неизвестная ошибка')
            
            # Советы по ошибкам
            error_advice = {
                'timeout': "Увеличить таймаут запросов или добавить повторные попытки.",
                '404': "Проверить актуальность URL.",
                '403': "Добавить обработку запрещенных запросов или изменить User-Agent."
            }
            
            advice = "Общие рекомендации: проверить доступность ресурса."
            for err_key, err_advice in error_advice.items():
                if err_key in error.lower():
                    advice = err_advice
                    break
            
            ws_summary.cell(row=row_num, column=10, value=f"Ошибка: {error}\nРекомендация: {advice}")
        
        ws_summary.cell(row=row_num, column=3, value="Успех" if result['status'] == 'success' else "Ошибка")
    
    # Форматирование
    for column in ws_summary.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_summary.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Настройка переноса текста для столбца с комментариями
    for row in ws_summary.iter_rows(min_row=2, max_col=10, max_row=len(results)+1):
        row[9].alignment = Alignment(wrapText=True, vertical='top')  # 10-й столбец (индекс 9)
    
    # === Лист 2: Детали сравнения ===
    ws_details = wb.create_sheet("Детали сравнения")
    
    # Заголовки
    headers = [
        "URL", "Примеры пропущенных строк", "Примеры добавленных строк", 
        "Анализ значимости", "Рекомендации"
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws_details.cell(row=1, column=col_num, value=header).font = Font(bold=True)
    
    # Данные
    for row_num, result in enumerate(results, 2):
        ws_details.cell(row=row_num, column=1, value=result['url'])
        
        if result['status'] == 'success':
            # Примеры пропущенных строк
            missing_examples = "\n".join([f"- {ex[:200]}{'...' if len(ex) > 200 else ''}" 
                                         for ex in result['example_missing']])
            ws_details.cell(row=row_num, column=2, value=missing_examples)
            
            # Примеры добавленных строк
            extra_examples = "\n".join([f"- {ex[:200]}{'...' if len(ex) > 200 else ''}" 
                                      for ex in result['example_extra']])
            ws_details.cell(row=row_num, column=3, value=extra_examples)
            
            # Анализ значимости
            analysis = []
            if result['missing_lines_count'] > 0:
                analysis.append(f"Пропущено {result['missing_lines_count']} строк. ")
                if any(re.search(r'(адрес|контакт|телефон|почта)', ex, re.I) for ex in result['example_missing']):
                    analysis.append("Пропущены контактные данные. ")
                if any(re.search(r'(вывод|заключение|итог)', ex, re.I) for ex in result['example_missing']):
                    analysis.append("Пропущены ключевые выводы. ")
            
            if result['extra_lines_count'] > 0:
                analysis.append(f"Добавлено {result['extra_lines_count']} строк. ")
                if any(re.search(r'(политик|экономик|конфликт)', ex, re.I) for ex in result['example_extra']):
                    analysis.append("Добавлен политизированный контент. ")
            
            ws_details.cell(row=row_num, column=4, value=" ".join(analysis) if analysis else "Незначительные расхождения")
            
            # Рекомендации
            recommendations = []
            if result['similarity'] < 50:
                recommendations.append("Требуется ручная проверка - низкое сходство.")
            elif result['similarity'] < 80:
                recommendations.append("Рекомендуется выборочная проверка.")
            
            if result['missing_lines_count'] > result['extra_lines_count']:
                recommendations.append("Проверить полноту извлечения контента.")
            elif result['extra_lines_count'] > result['missing_lines_count']:
                recommendations.append("Проверить на наличие нерелевантного контента.")
            
            ws_details.cell(row=row_num, column=5, value=" ".join(recommendations) if recommendations else "Приемлемый результат")
        else:
            ws_details.cell(row=row_num, column=2, value=result['error'])
    
    # Настройка переноса текста
    for row in ws_details.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True, vertical='top')
    
    # === Лист 3: Аналитика ===
    ws_analytics = wb.create_sheet("Аналитика")
    
    # Статистика
    stats = [
        ["Всего статей", len(results)],
        ["Успешно обработано", len([r for r in results if r['status'] == 'success'])],
        ["Ошибки обработки", len([r for r in results if r['status'] == 'error'])],
        ["Среднее сходство", np.mean([r['similarity'] for r in results if r['status'] == 'success']) if any(r['status'] == 'success' for r in results) else "N/A"],
        ["Максимальное сходство", max([r['similarity'] for r in results if r['status'] == 'success']) if any(r['status'] == 'success' for r in results) else "N/A"],
        ["Минимальное сходство", min([r['similarity'] for r in results if r['status'] == 'success']) if any(r['status'] == 'success' for r in results) else "N/A"]
    ]
    
    for row_num, stat in enumerate(stats, 1):
        ws_analytics.cell(row=row_num, column=1, value=stat[0]).font = Font(bold=True)
        ws_analytics.cell(row=row_num, column=2, value=stat[1])
    
    # Заключение
    conclusion = [
        "Заключение и рекомендации:",
        "",
        "1. Общая эффективность работы библиотеки:",
        f"- {'Высокая' if len([r for r in results if r['status'] == 'success' and r['similarity'] > 80]) / len(results) > 0.7 else 'Средняя' if len([r for r in results if r['status'] == 'success' and r['similarity'] > 50]) / len(results) > 0.7 else 'Низкая'} эффективность",
        "",
        "2. Основные проблемы:",
        "- Пропуск ключевых фрагментов текста" if any(r['status'] == 'success' and r['missing_lines_count'] > 10 for r in results) else "",
        "- Добавление нерелевантного контента" if any(r['status'] == 'success' and r['extra_lines_count'] > 10 for r in results) else "",
        "",
        "3. Рекомендации по улучшению:",
        "- Уточнить правила извлечения контента для уменьшения пропусков",
        "- Добавить проверку на релевантность извлекаемого контента",
        "- Реализовать обработку ошибок для проблемных URL",
        "",
        "4. Дополнительные предложения:",
        "- Внедрить ML-модель для оценки качества контента",
        "- Добавить систему оценки тональности текста",
        "- Реализовать сравнение с несколькими источниками"
    ]
    
    for row_num, line in enumerate(conclusion, len(stats) + 2):
        ws_analytics.cell(row=row_num, column=1, value=line)
    
    # Сохраняем файл
    wb.save(report_path)
    print(f"\n📊 Комплексный отчет сохранен: {report_path}")


def generate_improvement_tips(article):
    """Генерирует советы по улучшению для конкретной статьи"""
    tips = []
    
    if article['similarity'] < 50:
        tips.append("• Требуется серьезная доработка парсера для этого типа контента")
    elif article['similarity'] < 70:
        tips.append("• Рекомендуется оптимизация правил извлечения")
    
    if article['missing_lines_count'] > 10:
        tips.append("• Увеличить глубину анализа для уменьшения пропусков")
    
    if article['extra_lines_count'] > 5:
        tips.append("• Добавить фильтрацию нерелевантных блоков")
    
    if any('реклам' in ex.lower() for ex in article['example_extra']):
        tips.append("• Внедрить систему распознавания рекламного контента")
    
    if not tips:
        tips.append("• Качество извлечения удовлетворительное. Можно оптимизировать обработку специфичных элементов")
    
    return "<br>".join(tips)

def main():
    ensure_directory_exists()
    results = []

    print("=== Автоматический анализ текстов ===")
    print(f"Загружаем данные из файла: {INPUT_CSV}")

    try:
        df_input = pd.read_csv(INPUT_CSV, on_bad_lines='skip', sep=None, engine='python', encoding='utf-8-sig')
        df_input.columns = df_input.columns.str.strip()
        
        if not all(col.lower() in [c.lower() for c in df_input.columns] for col in ['URL', 'lib_text']):
            print("❌ Ошибка: Не найдены необходимые колонки!")
            return

        print(f"\n✅ Найдено {len(df_input)} записей. Начинаем обработку...\n")

        for idx, row in tqdm(df_input.iterrows(), total=len(df_input), desc="Обработка статей"):
            url = str(row['URL']).strip()
            lib_text = str(row['lib_text']).strip()

            if not lib_text:
                tqdm.write(f"⚠️ [{idx+1}] lib_text пустой. Пропускаем...")
                continue

            try:
                original_text = extract_text_from_url(url)
                
                if original_text.startswith("Ошибка"):
                    results.append({
                        'url': url,
                        'status': 'error',
                        'error': original_text
                    })
                    continue

                comparison = compare_texts(original_text, lib_text)
                results.append({
                    'url': url,
                    'similarity': comparison['similarity'],
                    'original_length': comparison['original_length'],
                    'lib_length': comparison['lib_length'],
                    'missing_lines_count': comparison['missing_lines_count'],
                    'extra_lines_count': comparison['extra_lines_count'],
                    'example_missing': comparison['example_missing'],
                    'example_extra': comparison['example_extra'],
                    'status': 'success'
                })

            except Exception as e:
                results.append({
                    'url': url,
                    'status': 'error',
                    'error': str(e)
                })
                tqdm.write(f"❌ [{idx+1}] Ошибка обработки {url}: {str(e)}")

        # Сохранение результатов
        save_results(results)
        generate_comprehensive_report(results)
        generate_html_report(results, OUTPUT_DIR)

    except Exception as e:
        print(f"❌ Не удалось обработать файл: {e}")
        return

    print(f"\n✅ Найдено {len(df_input)} записей. Начинаем обработку...\n")

    for idx, row in df_input.iterrows():
        url = str(row['URL']).strip()      # Обязательно с большой буквы!
        lib_text = str(row['lib_text']).strip()

        print(f"[{idx+1}/{len(df_input)}] Обработка: {url}")

        if not lib_text:
            print("⚠️ lib_text пустой. Пропускаем...\n")
            continue

        original_text = extract_text_from_url(url)

        if original_text.startswith("Ошибка"):
            print(f"❌ Ошибка парсинга: {original_text}\n")
            results.append({
                'url': url,
                'status': 'error',
                'error': original_text
            })
            continue

        comparison = compare_texts(original_text, lib_text)

        result = {
            'url': url,
            'similarity': comparison['similarity'],
            'original_length': comparison['original_length'],
            'lib_length': comparison['lib_length'],
            'missing_lines_count': comparison['missing_lines_count'],
            'extra_lines_count': comparison['extra_lines_count'],
            'example_missing': " | ".join(comparison['example_missing']),
            'example_extra': " | ".join(comparison['example_extra']),
            'status': 'success'
        }

        results.append(result)
        print(f"✅ Сходство: {comparison['similarity']}%\n")


if __name__ == "__main__":
    main()